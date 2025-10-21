from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from flask_wtf import FlaskForm
from wtforms import StringField, IntegerField, SelectField, TextAreaField, DateField, SubmitField, BooleanField
from wtforms.validators import DataRequired, Email, NumberRange
from datetime import datetime, date
import csv
import os
import secrets
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import tempfile

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///toy_planner.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'static/uploads'

db = SQLAlchemy(app)
migrate = Migrate(app, db)

# Create upload directory if it doesn't exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Models
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(100), nullable=False, unique=True)
    phone = db.Column(db.String(20), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships
    bookings = db.relationship('Booking', backref='user', lazy=True)

class Venue(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    district = db.Column(db.String(50), nullable=False)
    address = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    capacity_min = db.Column(db.Integer, nullable=False)
    capacity_max = db.Column(db.Integer, nullable=False)
    price_per_person = db.Column(db.Integer, nullable=False)  # in KZT
    phone = db.Column(db.String(20))
    email = db.Column(db.String(100))
    image_url = db.Column(db.String(200))
    event_types = db.Column(db.String(200))  # Comma-separated: wedding,corporate,tusau_keser
    
    # Relationships
    halls = db.relationship('Hall', backref='venue', lazy=True, cascade='all, delete-orphan')
    menu_items = db.relationship('MenuItem', backref='venue', lazy=True, cascade='all, delete-orphan')
    bookings = db.relationship('Booking', backref='venue', lazy=True)

class Hall(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    venue_id = db.Column(db.Integer, db.ForeignKey('venue.id'), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    capacity = db.Column(db.Integer, nullable=False)
    description = db.Column(db.Text)
    image_url = db.Column(db.String(200))

class MenuItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    venue_id = db.Column(db.Integer, db.ForeignKey('venue.id'), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    category = db.Column(db.String(50))  # appetizer, main, dessert, drink
    price = db.Column(db.Integer, nullable=False)  # in KZT
    description = db.Column(db.Text)

class Booking(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    venue_id = db.Column(db.Integer, db.ForeignKey('venue.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)  # Made nullable for backward compatibility
    client_name = db.Column(db.String(100), nullable=False)
    client_email = db.Column(db.String(100), nullable=False)
    client_phone = db.Column(db.String(20), nullable=False)
    event_type = db.Column(db.String(50), nullable=False)
    event_date = db.Column(db.Date, nullable=False)
    guest_count = db.Column(db.Integer, nullable=False)
    selected_hall_id = db.Column(db.Integer, db.ForeignKey('hall.id'))
    special_requests = db.Column(db.Text)
    total_amount = db.Column(db.Integer)  # in KZT
    deposit_paid = db.Column(db.Boolean, default=False)
    status = db.Column(db.String(20), default='pending')  # pending, confirmed, cancelled
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships
    selected_hall = db.relationship('Hall', backref='bookings')
    guests = db.relationship('Guest', backref='booking', lazy=True, cascade='all, delete-orphan')

class Guest(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    booking_id = db.Column(db.Integer, db.ForeignKey('booking.id'), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(100))
    phone = db.Column(db.String(20))
    rsvp_status = db.Column(db.String(20), default='pending')  # pending, attending, not_attending

class Feedback(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(100), nullable=False)
    feedback_type = db.Column(db.String(50), nullable=False)
    rating = db.Column(db.Integer, nullable=False)
    recommendation = db.Column(db.String(50), nullable=True)
    message = db.Column(db.Text, nullable=False)
    venue = db.Column(db.String(100))
    allow_contact = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    status = db.Column(db.String(20), default='new')  # new, reviewed, resolved

class Invitation(db.Model):
    """Model for event invitations with unique links"""
    id = db.Column(db.Integer, primary_key=True)
    booking_id = db.Column(db.Integer, db.ForeignKey('booking.id'), nullable=False)
    title = db.Column(db.String(200), nullable=False)
    message = db.Column(db.Text, nullable=False)
    event_time = db.Column(db.String(50))
    dress_code = db.Column(db.String(100))
    additional_info = db.Column(db.Text)
    unique_token = db.Column(db.String(100), unique=True, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    booking = db.relationship('Booking', backref='invitations')
    invited_guests = db.relationship('InvitedGuest', backref='invitation', lazy=True, cascade='all, delete-orphan')

class InvitedGuest(db.Model):
    """Model for individual guests invited through invitation link"""
    id = db.Column(db.Integer, primary_key=True)
    invitation_id = db.Column(db.Integer, db.ForeignKey('invitation.id'), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(100))
    phone = db.Column(db.String(20))
    plus_one = db.Column(db.Integer, default=0)
    rsvp_status = db.Column(db.String(20), default='pending')
    dietary_restrictions = db.Column(db.String(200))
    message_to_host = db.Column(db.Text)
    responded_at = db.Column(db.DateTime)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

# Forms
class VenueFilterForm(FlaskForm):
    event_type = SelectField('Event Type', choices=[
        ('', 'All Events'),
        ('wedding', 'Wedding'),
        ('kudalyk', 'Kudalyk'),
        ('betashar', 'Betashar'),
        ('tusau_keser', 'Tusau keser'),
        ('corporate_event', 'Corporate event'),
        ('birthday', 'Birthday'),
        ('anniversary', 'Anniversary'),
        ('graduation_prom', 'Graduation / Prom'),
        ('business_event', 'Business event'),
        ('national_cultural', 'National or cultural celebration')
    ])
    district = SelectField('District', choices=[
        ('', 'All Districts'),
        ('Bostandyk', 'Bostandyk'),
        ('Almaly', 'Almaly'),
        ('Auezov', 'Auezov'),
        ('Jetysu', 'Jetysu'),
        ('Medeu', 'Medeu'),
        ('Nauryzbay', 'Nauryzbay'),
        ('Turksib', 'Turksib'),
        ('Zhetysu', 'Zhetysu')
    ])
    guest_count = IntegerField('Number of Guests', validators=[NumberRange(min=1, max=1000)])
    max_price = IntegerField('Maximum Price per Person (KZT)', validators=[NumberRange(min=1000, max=50000)])
    submit = SubmitField('Search Venues')

class BookingForm(FlaskForm):
    client_name = StringField('Full Name', validators=[DataRequired()])
    client_email = StringField('Email', validators=[DataRequired(), Email()])
    client_phone = StringField('Phone Number', validators=[DataRequired()])
    event_type = SelectField('Event Type', choices=[], validators=[DataRequired()])
    event_date = DateField('Event Date', validators=[DataRequired()])
    guest_count = IntegerField('Number of Guests', validators=[DataRequired(), NumberRange(min=1, max=1000)])
    selected_hall_id = SelectField('Select Hall', coerce=int)
    special_requests = TextAreaField('Special Requests')
    submit = SubmitField('Continue to Payment')

class HostFilterForm(FlaskForm):
    language = StringField('Language')
    city = StringField('City')
    max_price = IntegerField('Max Price (KZT)', validators=[NumberRange(min=0, max=10000000)])
    submit = SubmitField('Search Hosts')

class MusicianFilterForm(FlaskForm):
    genre = StringField('Genre')
    city = StringField('City')
    max_price = IntegerField('Max Price (KZT)', validators=[NumberRange(min=0, max=10000000)])
    submit = SubmitField('Search Musicians')

class ProfileLookupForm(FlaskForm):
    email = StringField('Email Address', validators=[DataRequired(), Email()], 
                       render_kw={'placeholder': 'Enter your email to view your bookings'})
    submit = SubmitField('View My Profile')

class AddOnForm(FlaskForm):
    client_name = StringField('Full Name', validators=[DataRequired()])
    client_email = StringField('Email', validators=[DataRequired(), Email()])
    client_phone = StringField('Phone Number', validators=[DataRequired()])
    submit = SubmitField('Add to My Venue Booking')

class PaymentForm(FlaskForm):
    card_number = StringField('Card Number', validators=[DataRequired()], 
                             render_kw={'placeholder': '1234 5678 9012 3456', 'maxlength': '19'})
    card_holder = StringField('Cardholder Name', validators=[DataRequired()], 
                             render_kw={'placeholder': 'John Doe'})
    expiry_month = SelectField('Month', choices=[(str(i).zfill(2), str(i).zfill(2)) for i in range(1, 13)], validators=[DataRequired()])
    expiry_year = SelectField('Year', choices=[(str(i), str(i)) for i in range(2024, 2035)], validators=[DataRequired()])
    cvv = StringField('CVV', validators=[DataRequired()], 
                     render_kw={'placeholder': '123', 'maxlength': '4'})
    billing_address = TextAreaField('Billing Address', validators=[DataRequired()], 
                                   render_kw={'placeholder': 'Street address, City, Postal Code'})
    agree_terms = BooleanField('I agree to the Terms and Conditions', validators=[DataRequired()])
    submit = SubmitField('Complete Payment')

class InvitationForm(FlaskForm):
    """Form for creating event invitation"""
    title = StringField('Event Title', validators=[DataRequired()], 
                       render_kw={'placeholder': 'e.g., Wedding of John & Jane'})
    message = TextAreaField('Welcome Message', validators=[DataRequired()], 
                           render_kw={'placeholder': 'Write a warm message to your guests...', 'rows': 5})
    event_time = StringField('Event Time', validators=[DataRequired()], 
                            render_kw={'placeholder': 'e.g., 18:00'})
    dress_code = StringField('Dress Code', 
                            render_kw={'placeholder': 'e.g., Formal attire, Casual, Traditional'})
    additional_info = TextAreaField('Additional Information', 
                                   render_kw={'placeholder': 'Parking info, gift registry, etc.', 'rows': 3})
    submit = SubmitField('Create Invitation')

class GuestRSVPForm(FlaskForm):
    """Form for guests to RSVP"""
    name = StringField('Your Full Name', validators=[DataRequired()], 
                      render_kw={'placeholder': 'John Doe'})
    email = StringField('Email Address', validators=[Email()], 
                       render_kw={'placeholder': 'john@example.com'})
    phone = StringField('Phone Number', 
                       render_kw={'placeholder': '+7 700 123 4567'})
    plus_one = IntegerField('Number of Additional Guests', validators=[NumberRange(min=0, max=5)], 
                           default=0,
                           render_kw={'placeholder': '0'})
    rsvp_status = SelectField('Will you attend?', 
                             choices=[
                                 ('attending', '✓ Yes, I will attend'),
                                 ('not_attending', '✗ Sorry, I cannot attend')
                             ], 
                             validators=[DataRequired()])
    dietary_restrictions = StringField('Dietary Restrictions / Allergies', 
                                      render_kw={'placeholder': 'Vegetarian, Halal, Allergies, etc.'})
    message_to_host = TextAreaField('Message to Host (Optional)', 
                                   render_kw={'placeholder': 'Leave a message for the host...', 'rows': 3})
    submit = SubmitField('Submit RSVP')

# Helper functions
def find_or_create_user(name, email, phone):
    """Find existing user by email or create new one"""
    user = User.query.filter_by(email=email).first()
    if user:
        # Update user info if it has changed
        user.name = name
        user.phone = phone
        db.session.commit()
        return user
    else:
        # Create new user
        new_user = User(name=name, email=email, phone=phone)
        db.session.add(new_user)
        db.session.flush()  # Get the user ID
        return new_user

def load_csv_records(csv_filename):
    """Load generic records from a CSV file located in the instance folder.

    The function expects the first row to be headers and returns a list of
    dictionaries keyed by those headers. It tolerates missing files by
    returning an empty list and strips Byte Order Mark if present.
    """
    records = []
    try:
        # Try multiple possible paths
        possible_paths = [
            os.path.join('instance', csv_filename),
            os.path.join('PM_2 —final', 'instance', csv_filename),
            csv_filename
        ]
        
        instance_csv_path = None
        for path in possible_paths:
            if os.path.exists(path):
                instance_csv_path = path
                break
        
        if not instance_csv_path:
            return []
            
        with open(instance_csv_path, 'r', encoding='utf-8-sig', newline='') as csvfile:
            reader = csv.DictReader(csvfile)
            next_id = 1
            for row in reader:
                # Normalize keys: strip whitespace
                normalized = { (k.strip() if isinstance(k, str) else k): (v.strip() if isinstance(v, str) else v) for k, v in row.items() }
                # Assign sequential ID for detail pages
                if 'id' not in normalized or normalized.get('id', '') == '':
                    normalized['id'] = str(next_id)
                next_id += 1
                records.append(normalized)
    except Exception as e:
        print(f"Error loading CSV '{csv_filename}': {e}")
    return records

# Routes
@app.route('/')
def index():
    # Get featured venues for the main page
    venues = Venue.query.limit(6).all()  # Show first 6 venues
    return render_template('index.html', venues=venues)

@app.route('/venues')
def venues():
    form = VenueFilterForm()
    
    # Populate form with current filter values
    if request.args.get('event_type'):
        form.event_type.data = request.args.get('event_type')
    if request.args.get('district'):
        form.district.data = request.args.get('district')
    if request.args.get('guest_count'):
        form.guest_count.data = int(request.args.get('guest_count'))
    if request.args.get('max_price'):
        form.max_price.data = int(request.args.get('max_price'))
    
    query = Venue.query
    
    # Apply filters
    if request.args.get('event_type'):
        event_type = request.args.get('event_type')
        # Use LIKE with wildcards to match event types properly
        query = query.filter(Venue.event_types.like(f'%{event_type}%'))
    
    if request.args.get('district'):
        query = query.filter(Venue.district == request.args.get('district'))
    
    if request.args.get('guest_count'):
        guest_count = int(request.args.get('guest_count'))
        query = query.filter(Venue.capacity_min <= guest_count, Venue.capacity_max >= guest_count)
    
    if request.args.get('max_price'):
        max_price = int(request.args.get('max_price'))
        query = query.filter(Venue.price_per_person <= max_price)
    
    venues_list = query.all()
    return render_template('venues.html', venues=venues_list, form=form)

@app.route('/hosts')
def hosts():
    """Display list of event hosts from CSV with filters."""
    form = HostFilterForm()
    host_records = load_csv_records('hosts.csv')
    # Apply filters from query args
    language = request.args.get('language', '').strip()
    city = request.args.get('city', '').strip()
    max_price = request.args.get('max_price', '').strip()  # Debug info

    def parse_price(record):
        # Prefer per event else per hour
        for key in ['price_per_event', 'price_per_hour']:
            value = record.get(key)
            if value:
                try:
                    return int(value)
                except Exception:
                    continue
        return None

    filtered = []
    for r in host_records:
        if language and language.lower() not in (r.get('language', '') or '').lower():
            continue
        if city and city.lower() not in (r.get('city', '') or '').lower():
            continue
        if max_price:
            try:
                mp = int(max_price)
                price_val = parse_price(r)
                if price_val is not None and price_val > mp:
                    continue
            except Exception:
                pass
        filtered.append(r)

    return render_template('hosts.html', hosts=filtered, form=form)

@app.route('/musicians')
def musicians():
    """Display list of musicians/bands from CSV with filters."""
    form = MusicianFilterForm()
    musician_records = load_csv_records('musicians.csv')
    genre = request.args.get('genre', '').strip()
    city = request.args.get('city', '').strip()
    max_price = request.args.get('max_price', '').strip()  # Debug info

    def parse_price(record):
        for key in ['price_per_event', 'price_per_hour']:
            value = record.get(key)
            if value:
                try:
                    return int(value)
                except Exception:
                    continue
        return None

    filtered = []
    for r in musician_records:
        if genre and genre.lower() not in (r.get('genre', '') or '').lower():
            continue
        if city and city.lower() not in (r.get('city', '') or '').lower():
            continue
        if max_price:
            try:
                mp = int(max_price)
                price_val = parse_price(r)
                if price_val is not None and price_val > mp:
                    continue
            except Exception:
                pass
        filtered.append(r)

    return render_template('musicians.html', musicians=filtered, form=form)

@app.route('/host/<id>')
def host_detail(id):
    hosts = load_csv_records('hosts.csv')
    host = next((h for h in hosts if str(h.get('id')) == str(id)), None)
    if not host:
        return redirect(url_for('hosts'))
    return render_template('host_detail.html', host=host)

@app.route('/musician/<id>')
def musician_detail(id):
    musicians = load_csv_records('musicians.csv')
    artist = next((m for m in musicians if str(m.get('id')) == str(id)), None)
    if not artist:
        return redirect(url_for('musicians'))
    return render_template('musician_detail.html', musician=artist)

@app.route('/venue/<int:venue_id>')
def venue_detail(venue_id):
    venue = Venue.query.get_or_404(venue_id)
    return render_template('venue_detail.html', venue=venue)

@app.route('/book/<int:venue_id>', methods=['GET', 'POST'])
def book_venue(venue_id):
    venue = Venue.query.get_or_404(venue_id)
    form = BookingForm()
    
    # Populate hall choices
    if venue.halls:
        form.selected_hall_id.choices = [(hall.id, f"{hall.name} (Capacity: {hall.capacity})") 
                                         for hall in venue.halls]
    else:
        form.selected_hall_id.choices = [(0, "Main Hall (Default)")]
    
    # Populate event_type choices based on venue
    if venue.name == 'Navat Restaurant':
        form.event_type.choices = [
            ('corporate_event', 'Corporate dinners'),
            ('birthday', 'Birthday'),
            ('friendly_gatherings', 'Friendly gatherings / family dinners'),
            ('casual_celebrations', 'Casual celebrations')
        ]
    elif venue.name == 'Shyngyskhan Restaurant':
        form.event_type.choices = [
            ('wedding', 'Wedding'),
            ('kudalyk', 'Kudalyk'),
            ('betashar', 'Betashar'),
            ('tusau_keser', 'Tusau keser'),
            ('anniversary', 'Anniversary'),
            ('corporate_event', 'Corporate event (large)'),
            ('graduation_prom', 'Graduation / Prom')
        ]
    elif venue.name == 'Rixos Hotel Almaty':
        form.event_type.choices = [
            ('wedding', 'Wedding'),
            ('kudalyk', 'Kudalyk'),
            ('betashar', 'Betashar'),
            ('corporate_event', 'Corporate event'),
            ('anniversary', 'Anniversary'),
            ('graduation_prom', 'Graduation / Prom')
        ]
    else:
        # Default choices for other venues
        form.event_type.choices = [
            ('wedding', 'Wedding'),
            ('corporate_event', 'Corporate event'),
            ('birthday', 'Birthday'),
            ('anniversary', 'Anniversary')
        ]
    
    # Populate event_type with URL parameter if available
    if request.args.get('event_type'):
        form.event_type.data = request.args.get('event_type')
    
    if form.validate_on_submit():
        # Store booking data in session for payment confirmation
        from flask import session
        session['booking_data'] = {
            'venue_id': venue_id,
            'client_name': form.client_name.data,
            'client_email': form.client_email.data,
            'client_phone': form.client_phone.data,
            'event_type': form.event_type.data,
            'event_date': form.event_date.data.isoformat(),
            'guest_count': form.guest_count.data,
            'selected_hall_id': form.selected_hall_id.data,
            'special_requests': form.special_requests.data,
            'total_amount': venue.price_per_person * form.guest_count.data
        }
        
        return redirect(url_for('payment_confirmation'))
    
    return render_template('book_venue.html', venue=venue, form=form)

@app.route('/payment/confirmation', methods=['GET', 'POST'])
def payment_confirmation():
    from flask import session
    from datetime import datetime as dt
    
    # Check if booking data exists in session
    if 'booking_data' not in session:
        flash('No booking data found. Please start your booking again.', 'error')
        return redirect(url_for('venues'))
    
    booking_data = session['booking_data']
    venue = Venue.query.get_or_404(booking_data['venue_id'])
    selected_hall = Hall.query.get(booking_data['selected_hall_id']) if booking_data['selected_hall_id'] else None
    
    form = PaymentForm()
    
    if form.validate_on_submit():
        # Create the actual booking after payment
        user = find_or_create_user(
            name=booking_data['client_name'],
            email=booking_data['client_email'],
            phone=booking_data['client_phone']
        )
        
        booking = Booking(
            venue_id=booking_data['venue_id'],
            user_id=user.id,
            client_name=booking_data['client_name'],
            client_email=booking_data['client_email'],
            client_phone=booking_data['client_phone'],
            event_type=booking_data['event_type'],
            event_date=dt.fromisoformat(booking_data['event_date']).date(),
            guest_count=booking_data['guest_count'],
            selected_hall_id=booking_data['selected_hall_id'] if booking_data['selected_hall_id'] != 0 else None,
            special_requests=booking_data['special_requests'],
            total_amount=booking_data['total_amount'],
            deposit_paid=True,
            status='confirmed'
        )
        
        db.session.add(booking)
        db.session.commit()
        
        # Clear session data
        session.pop('booking_data', None)
        
        flash('Payment successful! Your booking has been confirmed.', 'success')
        return redirect(url_for('booking_confirmation', booking_id=booking.id))
    
    return render_template('payment_confirmation.html', 
                         booking_data=booking_data, 
                         venue=venue, 
                         selected_hall=selected_hall,
                         form=form)

@app.route('/booking/<int:booking_id>/confirmation')
def booking_confirmation(booking_id):
    booking = Booking.query.get_or_404(booking_id)
    return render_template('booking_confirmation.html', booking=booking)

@app.route('/rsvp/<int:booking_id>')
def guest_rsvp(booking_id):
    booking = Booking.query.get_or_404(booking_id)
    return render_template('guest_rsvp.html', booking=booking)

@app.route('/api/rsvp', methods=['POST'])
def submit_rsvp():
    data = request.json
    guest = Guest(
        booking_id=data['booking_id'],
        name=data['name'],
        email=data.get('email', ''),
        phone=data.get('phone', ''),
        rsvp_status=data['status']
    )
    
    db.session.add(guest)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'RSVP submitted successfully!'})

@app.route('/feedback')
def feedback_page():
    return render_template('feedback.html')

@app.route('/submit_feedback', methods=['POST'])
def submit_feedback():
    try:
        # Get form data with validation
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip()
        feedback_type = request.form.get('feedback_type', '').strip()
        rating_str = request.form.get('rating', '').strip()
        recommendation = request.form.get('recommendation', '').strip()
        message = request.form.get('message', '').strip()
        venue = request.form.get('venue', '').strip()
        # Removed allow_contact checkbox as per user request
        
        # Validate required fields
        if not all([name, email, feedback_type, rating_str, recommendation, message]):
            flash('Please fill in all required fields.', 'error')
            return redirect(url_for('feedback_page'))
        
        # Validate rating
        try:
            rating = int(rating_str)
            if rating < 1 or rating > 5:
                raise ValueError("Invalid rating")
        except ValueError:
            flash('Please select a valid rating.', 'error')
            return redirect(url_for('feedback_page'))
        
        # Create feedback record
        feedback = Feedback(
            name=name,
            email=email,
            feedback_type=feedback_type,
            rating=rating,
            recommendation=recommendation,
            message=message,
            venue=venue if venue else None,
            allow_contact=False  # Always set to False since checkbox is removed
        )
        
        db.session.add(feedback)
        db.session.commit()
        
        # Save feedback data to Excel file AFTER database commit so feedback.id is available
        try:
            save_feedback_to_excel(feedback)
            print(f"Successfully saved feedback ID {feedback.id} to Excel")
        except Exception as excel_error:
            print(f"Warning: Could not save feedback ID {feedback.id} to Excel: {excel_error}")
            # Continue with success message even if Excel save fails
        
        flash('Thank you for your feedback! We appreciate your input and will use it to improve our services.', 'success')
        return redirect(url_for('feedback_success'))
        
    except Exception as e:
        print(f"Feedback submission error: {e}")  # Debug logging
        db.session.rollback()
        flash('There was an error submitting your feedback. Please try again.', 'error')
        return redirect(url_for('feedback_page'))

@app.route('/feedback/success')
def feedback_success():
    return render_template('feedback_success.html')

def save_feedback_to_excel(feedback):
    """Save feedback data to Excel file"""
    try:
        excel_file_path = os.path.join('static', 'feedback_data.xlsx')
        
        # Check if file exists
        if os.path.exists(excel_file_path):
            # Load existing workbook
            from openpyxl import load_workbook
            wb = load_workbook(excel_file_path)
            ws = wb.active
        else:
            # Create new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Feedback Data"
            
            # Create headers with styling
            headers = [
                'ID', 'Date & Time', 'Name', 'Email', 'Overall Rating', 
                'Feedback Type', 'Recommendation', 'Related Venue', 'Message', 'Status'
            ]
            
            # Style headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Set column widths
            column_widths = [8, 20, 25, 30, 15, 15, 20, 25, 50, 12]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        
        # Convert rating number to text
        rating_text = {
            5: "😍 Excellent",
            4: "😊 Very Good", 
            3: "😐 Good",
            2: "😕 Fair",
            1: "😞 Poor"
        }.get(feedback.rating, f"Rating {feedback.rating}")
        
        # Convert feedback type to readable format
        feedback_type_text = {
            'compliment': '👏 Compliment',
            'suggestion': '💡 Suggestion',
            'complaint': '😟 Complaint',
            'general': '💬 General'
        }.get(feedback.feedback_type, feedback.feedback_type.title())
        
        # Convert recommendation to readable format
        recommendation_text = {
            'definitely': 'Definitely! 🎉',
            'probably': 'Probably 👍',
            'maybe': 'Maybe 🤔',
            'probably_not': 'Probably not 👎',
            'definitely_not': 'Definitely not ❌'
        }.get(feedback.recommendation, feedback.recommendation)
        
        # Add new row
        new_row = [
            feedback.id,
            feedback.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            feedback.name,
            feedback.email,
            rating_text,
            feedback_type_text,
            recommendation_text,
            feedback.venue or 'N/A',
            feedback.message,
            feedback.status.title()
        ]
        
        ws.append(new_row)
        
        # Style the new row
        row_num = ws.max_row
        for col in range(1, len(new_row) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            # Alternate row colors
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        # Save the workbook
        wb.save(excel_file_path)
        print(f"Feedback saved to Excel: {excel_file_path}")
        
    except Exception as e:
        print(f"Error saving feedback to Excel: {e}")

@app.route('/download_feedback_excel')
def download_feedback_excel():
    """Download the feedback Excel file"""
    try:
        excel_file_path = os.path.join('static', 'feedback_data.xlsx')
        
        if os.path.exists(excel_file_path):
            return send_file(
                excel_file_path,
                as_attachment=True,
                download_name=f'toy_planner_feedback_{datetime.now().strftime("%Y%m%d")}.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            flash('No feedback data available for download.', 'info')
            return redirect(url_for('index'))
            
    except Exception as e:
        print(f"Error downloading feedback Excel: {e}")
        flash('Error downloading feedback data.', 'error')
        return redirect(url_for('index'))

@app.route('/profile', methods=['GET', 'POST'])
def profile():
    """Profile page where users can enter email and phone to view their bookings"""
    form = ProfileLookupForm()
    
    if form.validate_on_submit():
        email = form.email.data
        phone = request.form.get('phone', '').strip()
        
        # Verify both email AND phone match
        user = User.query.filter_by(email=email, phone=phone).first()
        
        if user:
            return redirect(url_for('user_profile', user_id=user.id))
        else:
            flash('Email and phone combination not found. Please check your credentials or make a booking first.', 'warning')
    
    return render_template('profile.html', form=form)

@app.route('/profile/<int:user_id>')
def user_profile(user_id):
    """Display user's booking history and profile information"""
    user = User.query.get_or_404(user_id)
    
    # Get all bookings for this user, ordered by most recent first
    bookings = Booking.query.filter_by(user_id=user_id).order_by(Booking.created_at.desc()).all()
    
    # Separate bookings by status
    upcoming_bookings = [b for b in bookings if b.event_date >= date.today() and b.status in ['pending', 'confirmed']]
    past_bookings = [b for b in bookings if b.event_date < date.today() or b.status == 'cancelled']
    
    # Calculate total spent (30% deposits only - what user actually paid)
    total_spent = sum(int(booking.total_amount * 0.3) for booking in bookings if booking.total_amount and booking.deposit_paid)
    
    return render_template('user_profile.html', 
                         user=user, 
                         upcoming_bookings=upcoming_bookings, 
                         past_bookings=past_bookings,
                         total_spent=total_spent)

def _parse_price_from_record(record):
    for key in ['price_per_event', 'price_per_hour']:
        value = record.get(key)
        if value:
            try:
                return int(value)
            except Exception:
                continue
    return None

def _find_latest_relevant_booking_by_email(email):
    # Prefer upcoming bookings first, else fall back to latest any booking
    upcoming = (Booking.query
                .filter(Booking.client_email == email, Booking.event_date >= date.today())
                .order_by(Booking.event_date.asc(), Booking.created_at.desc())
                .first())
    if upcoming:
        return upcoming
    return (Booking.query
            .filter(Booking.client_email == email)
            .order_by(Booking.created_at.desc())
            .first())

@app.route('/book_host/<id>', methods=['GET', 'POST'])
def book_host(id):
    hosts = load_csv_records('hosts.csv')
    host = next((h for h in hosts if str(h.get('id')) == str(id)), None)
    if not host:
        flash('Selected host not found.', 'error')
        return redirect(url_for('hosts'))

    form = AddOnForm()
    if form.validate_on_submit():
        user = find_or_create_user(
            name=form.client_name.data,
            email=form.client_email.data,
            phone=form.client_phone.data
        )

        booking = _find_latest_relevant_booking_by_email(form.client_email.data)
        if not booking:
            flash('No venue booking found for this email. Please book a venue first.', 'info')
            return redirect(url_for('venues'))

        add_price = _parse_price_from_record(host) or 0
        booking.user_id = user.id
        booking.total_amount = (booking.total_amount or 0) + add_price
        note_parts = []
        if booking.special_requests:
            note_parts.append(booking.special_requests)
        note_parts.append(f"Added Host: {host.get('name', 'Host')} (+{add_price} KZT)")
        booking.special_requests = '\n'.join([p for p in note_parts if p])
        db.session.commit()

        flash('Host has been added to your venue booking and total updated.', 'success')
        return redirect(url_for('user_profile', user_id=user.id))

    return render_template('book_host.html', host=host, form=form)

@app.route('/book_musician/<id>', methods=['GET', 'POST'])
def book_musician(id):
    musicians = load_csv_records('musicians.csv')
    artist = next((m for m in musicians if str(m.get('id')) == str(id)), None)
    if not artist:
        flash('Selected musician not found.', 'error')
        return redirect(url_for('musicians'))

    form = AddOnForm()
    if form.validate_on_submit():
        user = find_or_create_user(
            name=form.client_name.data,
            email=form.client_email.data,
            phone=form.client_phone.data
        )

        booking = _find_latest_relevant_booking_by_email(form.client_email.data)
        if not booking:
            flash('No venue booking found for this email. Please book a venue first.', 'info')
            return redirect(url_for('venues'))

        add_price = _parse_price_from_record(artist) or 0
        booking.user_id = user.id
        booking.total_amount = (booking.total_amount or 0) + add_price
        note_parts = []
        if booking.special_requests:
            note_parts.append(booking.special_requests)
        note_parts.append(f"Added Musician: {artist.get('name', 'Artist')} (+{add_price} KZT)")
        booking.special_requests = '\n'.join([p for p in note_parts if p])
        db.session.commit()

        flash('Musician has been added to your venue booking and total updated.', 'success')
        return redirect(url_for('user_profile', user_id=user.id))

    return render_template('book_musician.html', musician=artist, form=form)

# Initialize database
def create_tables():
    with app.app_context():
        db.create_all()
        
        # Add sample data if database is empty
        if Venue.query.count() == 0:
            add_sample_data()

def add_sample_data():
    # Sample venues
    venues_data = [
        {
            'name': 'Grand Almaty Hotel',
            'district': 'Bostandyk',
            'address': 'Al-Farabi Avenue 77, Almaty',
            'description': 'Elegant venue perfect for weddings and corporate events',
            'capacity_min': 50,
            'capacity_max': 300,
            'price_per_person': 12000,
            'phone': '+7 727 291 4747',
            'email': 'events@grandalmaty.kz',
            'event_types': 'wedding,corporate,tusau_keser'
        },
        {
            'name': 'Kok-Tobe Restaurant',
            'district': 'Medeu',
            'address': 'Kok-Tobe Hill, Almaty',
            'description': 'Stunning mountain views for memorable celebrations',
            'capacity_min': 30,
            'capacity_max': 200,
            'price_per_person': 15000,
            'phone': '+7 727 273 5555',
            'email': 'info@koktobe.kz',
            'event_types': 'wedding,tusau_keser,kudalyk'
        },
        {
            'name': 'Atakent Palace',
            'district': 'Almaly',
            'address': 'Atakent Exhibition Center, Almaty',
            'description': 'Large conference halls ideal for corporate events',
            'capacity_min': 100,
            'capacity_max': 500,
            'price_per_person': 8000,
            'phone': '+7 727 378 7878',
            'email': 'bookings@atakent.kz',
            'event_types': 'corporate,wedding'
        },
        {
            'name': 'Navat Restaurant',
            'district': 'Almaly',
            'address': 'Abay Avenue 150, Almaty',
            'description': 'Experience authentic Kazakh cuisine with traditional dishes and modern presentation. Perfect for family gatherings and celebrations.',
            'capacity_min': 50,
            'capacity_max': 150,
            'price_per_person': 2700,
            'phone': '+7 727 123 4567',
            'email': 'info@navat.kz',
            'image_url': 'images/NAvat.png',
                'event_types': 'corporate_event,birthday,friendly_gatherings,casual_celebrations'
        },
        {
            'name': 'Shyngyskhan Restaurant',
            'district': 'Bostandyk',
            'address': 'Dostyk Avenue 85, Almaty',
            'description': 'Premium dining experience with luxurious banquet options and comprehensive event packages. Perfect for weddings and special celebrations.',
            'capacity_min': 50,
            'capacity_max': 300,
            'price_per_person': 16000,
            'phone': '+7 727 234 5678',
            'email': 'events@shyngyskhan.kz',
            'image_url': 'images/Shyngyskhan.png.webp',
            'event_types': 'wedding,kudalyk,betashar,tusau_keser,anniversary,corporate_event,graduation_prom'
        },
        {
            'name': 'Rixos Hotel Almaty',
            'district': 'Bostandyk',
            'address': 'Abay Avenue 145/1, Almaty',
            'description': 'Luxury hotel restaurant offering international cuisine with premium dining experience. Perfect for upscale events, corporate meetings, and special celebrations.',
            'capacity_min': 30,
            'capacity_max': 200,
            'price_per_person': 4000,
            'phone': '+7 727 377 7777',
            'email': 'restaurant@rixos-almaty.kz',
            'image_url': 'images/Rixos.jpeg',
            'event_types': 'wedding,kudalyk,betashar,corporate_event,anniversary,graduation_prom'
        }
    ]
    
    for venue_data in venues_data:
        venue = Venue(**venue_data)
        db.session.add(venue)
        db.session.flush()  # Get the venue ID
        
        # Add sample halls
        if venue.name == 'Grand Almaty Hotel':
            halls = [
                {'name': 'Small Banquet Hall', 'capacity': 80, 'description': 'Intimate setting for smaller gatherings'},
                {'name': 'Grand Ballroom', 'capacity': 250, 'description': 'Spacious hall for large celebrations'}
            ]
        elif venue.name == 'Kok-Tobe Restaurant':
            halls = [
                {'name': 'Terrace Hall', 'capacity': 120, 'description': 'Open-air dining with city views'},
                {'name': 'Mountain View Hall', 'capacity': 180, 'description': 'Indoor hall with panoramic windows'}
            ]
        elif venue.name == 'Navat Restaurant':
            halls = [
                {'name': 'Main Dining Hall', 'capacity': 80, 'description': 'Traditional Kazakh dining experience'},
                {'name': 'Private Banquet Room', 'capacity': 120, 'description': 'Intimate space for special occasions'}
            ]
        elif venue.name == 'Shyngyskhan Restaurant':
            halls = [
                {'name': 'Premium Hall', 'capacity': 150, 'description': 'Luxurious dining with premium amenities'},
                {'name': 'Grand Banquet Hall', 'capacity': 250, 'description': 'Spacious hall for large celebrations'}
            ]
        elif venue.name == 'Rixos Hotel Almaty':
            halls = [
                {'name': 'Main Restaurant Hall', 'capacity': 80, 'description': 'Elegant dining room with international cuisine'},
                {'name': 'Private Dining Room', 'capacity': 40, 'description': 'Intimate space for special occasions'},
                {'name': 'Terrace Restaurant', 'capacity': 120, 'description': 'Outdoor dining with city views'}
            ]
        else:
            halls = [
                {'name': 'Conference Hall A', 'capacity': 200, 'description': 'Modern facilities for business events'},
                {'name': 'Grand Assembly Hall', 'capacity': 400, 'description': 'Large space for major events'}
            ]
        
        for hall_data in halls:
            hall = Hall(venue_id=venue.id, **hall_data)
            db.session.add(hall)
        
        # Add sample menu items
        if venue.name == 'Navat Restaurant':
            menu_items = [
                {'name': 'Beshbarmak with Lamb', 'category': 'main', 'price': 4000, 'description': 'Traditional Kazakh dish with lamb (600g)'},
                {'name': 'Beshbarmak with Beef', 'category': 'main', 'price': 4000, 'description': 'Traditional Kazakh dish with beef (540g)'},
                {'name': 'Beshbarmak with Horse Meat', 'category': 'main', 'price': 4700, 'description': 'Traditional Kazakh dish with horse meat (600g)'},
                {'name': 'Beef Stroganoff with Rice', 'category': 'main', 'price': 3800, 'description': 'Classic stroganoff served with rice (320g)'},
                {'name': 'Pilaf Toy Ashi', 'category': 'main', 'price': 3100, 'description': 'Traditional rice pilaf (450g)'},
                {'name': 'Lamb Rib Kebab', 'category': 'main', 'price': 3800, 'description': 'Grilled lamb ribs (250g)'},
                {'name': 'Manti with Meat', 'category': 'main', 'price': 3400, 'description': 'Traditional dumplings with meat (450g)'},
                {'name': 'Ash-Lyam-Fu', 'category': 'main', 'price': 2700, 'description': 'Traditional cold noodle dish (650g)'},
                {'name': 'Coca-Cola Glass', 'category': 'drink', 'price': 1200, 'description': 'Refreshing cola (250ml)'},
                {'name': 'Ice Tea Orange', 'category': 'drink', 'price': 1800, 'description': 'Orange flavored iced tea (500ml)'}
            ]
        elif venue.name == 'Shyngyskhan Restaurant':
            menu_items = [
                # Menu 1: Premium All Inclusive Package (from ShyngysKhanDB.csv)
                {'name': 'Premium All Inclusive Package', 'category': 'premium_package', 'price': 25000, 'description': 'Complete premium package per person'},
                {'name': 'Musical and lighting equipment and LED screen', 'category': 'premium_package', 'price': 200000, 'description': 'Included in premium package'},
                {'name': 'Buffet (canapés, fruit assortment)', 'category': 'premium_package', 'price': 0, 'description': 'Gift - included in premium package'},
                {'name': 'New Year\'s hall decoration', 'category': 'premium_package', 'price': 0, 'description': 'Gift - included in premium package'},
                {'name': 'Bar menu for each table (12 people)', 'category': 'premium_package', 'price': 0, 'description': 'Unlimited - included in premium package'},
                {'name': 'Premium Vodka \'Zerna Severa\'', 'category': 'premium_package', 'price': 0, 'description': 'Included in premium package'},
                {'name': 'Georgian Red Wine \'Kinzmarauli\'', 'category': 'premium_package', 'price': 0, 'description': 'Included in premium package'},
                {'name': 'Italian White Wine \'Ribolla Gialla\'', 'category': 'premium_package', 'price': 0, 'description': 'Included in premium package'},
                {'name': 'Italian Champagne', 'category': 'premium_package', 'price': 0, 'description': 'Included in premium package'},
                {'name': 'Morse Compote', 'category': 'premium_package', 'price': 0, 'description': 'Included in premium package'},
                {'name': 'Water', 'category': 'premium_package', 'price': 0, 'description': 'Included in premium package'},
                {'name': 'Kona', 'category': 'premium_package', 'price': 0, 'description': 'Included in premium package'},
                {'name': 'Tea assortment', 'category': 'premium_package', 'price': 0, 'description': 'Unlimited - included in premium package'},
                {'name': 'New Year\'s snacks', 'category': 'premium_package', 'price': 0, 'description': '6 types to choose from - included in premium package'},
                {'name': 'Horse Assortment', 'category': 'premium_package', 'price': 0, 'description': 'Included in premium package'},
                {'name': 'Fish Assortment', 'category': 'premium_package', 'price': 0, 'description': 'Included in premium package'},
                {'name': 'Fresh Vegetable Bouquet', 'category': 'premium_package', 'price': 0, 'description': 'Included in premium package'},
                {'name': 'Assorted Pickles', 'category': 'premium_package', 'price': 0, 'description': 'Included in premium package'},
                {'name': 'Russian Snack', 'category': 'premium_package', 'price': 0, 'description': 'Included in premium package'},
                {'name': 'Eggplant Rolls', 'category': 'premium_package', 'price': 0, 'description': 'Included in premium package'},
                {'name': 'Oriental Samsa with Chicken', 'category': 'premium_package', 'price': 0, 'description': 'Included in premium package'},
                {'name': 'Vegetable Spring Rolls', 'category': 'premium_package', 'price': 0, 'description': 'Included in premium package'},
                
                # Menu 2: Banquet Menu Package (from ShyngysKhan1DB.csv)
                {'name': 'Banquet Menu Package', 'category': 'banquet_package', 'price': 16000, 'description': 'Complete banquet menu per person'},
                {'name': 'Cold Appetizers', 'category': 'banquet_package', 'price': 0, 'description': '3 types to choose from - included in banquet menu'},
                {'name': 'Horse Assortment (goat meat, jerky, lettuce)', 'category': 'banquet_package', 'price': 0, 'description': 'Included in banquet menu'},
                {'name': 'Fish Assortment (smoked salmon, smoked balyk, lemon, lettuce, olives)', 'category': 'banquet_package', 'price': 0, 'description': 'Included in banquet menu'},
                {'name': 'Fresh Vegetable Bouquet (cucumbers, tomatoes, bell peppers, brynza, olives, herbs)', 'category': 'banquet_package', 'price': 0, 'description': 'Included in banquet menu'},
                {'name': 'Assorted Pickles (pickled cucumbers, pickled cherry tomatoes, pickled pattypans, pickled cabbage \'Petal\', honey mushrooms)', 'category': 'banquet_package', 'price': 0, 'description': 'Included in banquet menu'},
                {'name': 'Russian Snack (lightly salted herring, boiled potatoes, gherkins, fermented cabbage, green onions)', 'category': 'banquet_package', 'price': 0, 'description': 'Included in banquet menu'},
                {'name': 'Eggplant Rolls (eggplant with feta cheese and tomatoes)', 'category': 'banquet_package', 'price': 0, 'description': 'Included in banquet menu'},
                {'name': 'Hot Appetizers', 'category': 'banquet_package', 'price': 0, 'description': '1 type to choose from - included in banquet menu'},
                {'name': 'Chicken Wings in Teriyaki Sauce', 'category': 'banquet_package', 'price': 0, 'description': 'Included in banquet menu'},
                {'name': 'Chicken Wings in BBQ Sauce', 'category': 'banquet_package', 'price': 0, 'description': 'Included in banquet menu'},
                {'name': 'Oriental Samsa with Chicken', 'category': 'banquet_package', 'price': 0, 'description': 'Included in banquet menu'},
                {'name': 'Salads', 'category': 'banquet_package', 'price': 0, 'description': '4 types to choose from - included in banquet menu'},
                {'name': 'Caesar with Chicken (iceberg lettuce, cherry tomatoes, quail eggs, toasted bread, Caesar dressing, parmesan)', 'category': 'banquet_package', 'price': 0, 'description': 'Included in banquet menu'},
                {'name': 'Crispy Eggplants (tomatoes, battered eggplants, sesame and Sweet Chili & Kikkoman sauce)', 'category': 'banquet_package', 'price': 0, 'description': 'Included in banquet menu'}
            ]
        elif venue.name == 'Rixos Hotel Almaty':
            menu_items = [
                # Salads
                {'name': 'Mixed Green Salad', 'category': 'salad', 'price': 4000, 'description': 'Fresh mixed greens with seasonal vegetables'},
                {'name': 'Burrata Salad', 'category': 'salad', 'price': 7500, 'description': 'Italian burrata cheese with fresh tomatoes and basil'},
                {'name': 'Red Tuna Tartar', 'category': 'salad', 'price': 7000, 'description': 'Fresh red tuna tartar with avocado and sesame'},
                
                # Main Courses - Grilled Menu
                {'name': 'Grilled Menu', 'category': 'main', 'price': 11000, 'description': 'Complete grilled menu selection'},
                {'name': 'Beef Tenderloin 200g', 'category': 'main', 'price': 25000, 'description': 'Premium beef tenderloin grilled to perfection'},
                {'name': 'Rib Eye Steak 400g', 'category': 'main', 'price': 11000, 'description': 'Juicy rib eye steak with your choice of sides'},
                {'name': 'Lamb Chops 250g', 'category': 'main', 'price': 6000, 'description': 'Tender lamb chops with herbs and spices'},
                {'name': 'Chicken Breast 200g', 'category': 'main', 'price': 6500, 'description': 'Grilled chicken breast with seasonal vegetables'},
                {'name': 'Turkish Meatballs 200g', 'category': 'main', 'price': 11000, 'description': 'Traditional Turkish meatballs with sauce'},
                {'name': 'Beef Kebab 200g', 'category': 'main', 'price': 6000, 'description': 'Tender beef kebab with fresh vegetables'},
                {'name': 'Chicken Kebab 200g', 'category': 'main', 'price': 6000, 'description': 'Marinated chicken kebab with herbs'},
                
                # Desserts
                {'name': 'Chocolate Mousse with Cherry', 'category': 'dessert', 'price': 4500, 'description': 'Rich chocolate mousse topped with fresh cherries'},
                {'name': 'Pistachio Roll', 'category': 'dessert', 'price': 4500, 'description': 'Delicate pistachio cream roll with pistachio nuts'},
                {'name': 'Profiteroles with Caramel', 'category': 'dessert', 'price': 4500, 'description': 'Light choux pastry filled with cream and caramel sauce'},
                {'name': 'Apple Strudel with Vanilla Ice Cream', 'category': 'dessert', 'price': 4500, 'description': 'Traditional apple strudel served with vanilla ice cream'},
                {'name': 'Seasonal Fruits', 'category': 'dessert', 'price': 7000, 'description': 'Fresh seasonal fruit platter'},
                {'name': 'Cheese Platter with Dried Fruits', 'category': 'dessert', 'price': 15000, 'description': 'Selection of premium cheeses with dried fruits and nuts'},
                {'name': 'Ice Cream (1 Scoop)', 'category': 'dessert', 'price': 2000, 'description': 'Single scoop of premium ice cream'},
                
                # Beverages - Water
                {'name': 'Tassay Water Still (Small)', 'category': 'drink', 'price': 1500, 'description': 'Premium still water (0.5L)'},
                {'name': 'Tassay Water Still (Large)', 'category': 'drink', 'price': 2000, 'description': 'Premium still water (1L)'},
                {'name': 'Tassay Water Sparkling (Small)', 'category': 'drink', 'price': 1500, 'description': 'Premium sparkling water (0.5L)'},
                {'name': 'Tassay Water Sparkling (Large)', 'category': 'drink', 'price': 2000, 'description': 'Premium sparkling water (1L)'},
                {'name': 'Tassay Excellent', 'category': 'drink', 'price': 3000, 'description': 'Premium mineral water'},
                {'name': 'Tassay Emerald', 'category': 'drink', 'price': 3000, 'description': 'Premium mineral water'},
                {'name': 'Borjomi', 'category': 'drink', 'price': 3000, 'description': 'Georgian mineral water'},
                {'name': 'Perrier', 'category': 'drink', 'price': 3500, 'description': 'French sparkling mineral water'},
                {'name': 'San Pellegrino (Small)', 'category': 'drink', 'price': 3000, 'description': 'Italian sparkling water (0.5L)'},
                {'name': 'San Pellegrino (Large)', 'category': 'drink', 'price': 5000, 'description': 'Italian sparkling water (1L)'},
                {'name': 'Acqua Panna (Small)', 'category': 'drink', 'price': 3000, 'description': 'Italian still water (0.5L)'},
                {'name': 'Acqua Panna (Large)', 'category': 'drink', 'price': 5000, 'description': 'Italian still water (1L)'}
            ]
        else:
            menu_items = [
                {'name': 'Beshbarmak', 'category': 'main', 'price': 3500, 'description': 'Traditional Kazakh dish'},
                {'name': 'Pilaf', 'category': 'main', 'price': 2800, 'description': 'Aromatic rice with meat'},
                {'name': 'Baursak', 'category': 'appetizer', 'price': 1200, 'description': 'Traditional fried bread'},
                {'name': 'Shubat', 'category': 'drink', 'price': 800, 'description': 'Fermented camel milk'},
                {'name': 'Wedding Cake', 'category': 'dessert', 'price': 15000, 'description': 'Custom wedding cake (per cake)'}
            ]
        
        for item_data in menu_items:
            menu_item = MenuItem(venue_id=venue.id, **item_data)
            db.session.add(menu_item)
    
    db.session.commit()

# ========== INVITATION SYSTEM ROUTES ==========

@app.route('/booking/<int:booking_id>/create-invitation', methods=['GET', 'POST'])
def create_invitation(booking_id):
    """Create invitation for a booking"""
    booking = Booking.query.get_or_404(booking_id)
    existing_invitation = Invitation.query.filter_by(booking_id=booking_id).first()
    
    form = InvitationForm()
    
    if form.validate_on_submit():
        if existing_invitation:
            existing_invitation.title = form.title.data
            existing_invitation.message = form.message.data
            existing_invitation.event_time = form.event_time.data
            existing_invitation.dress_code = form.dress_code.data
            existing_invitation.additional_info = form.additional_info.data
            invitation = existing_invitation
        else:
            invitation = Invitation(
                booking_id=booking_id,
                title=form.title.data,
                message=form.message.data,
                event_time=form.event_time.data,
                dress_code=form.dress_code.data,
                additional_info=form.additional_info.data,
                unique_token=secrets.token_urlsafe(32)
            )
            db.session.add(invitation)
        
        db.session.commit()
        flash('Invitation created successfully! Share the link with your guests.', 'success')
        return redirect(url_for('invitation_preview', token=invitation.unique_token))
    
    if existing_invitation:
        form.title.data = existing_invitation.title
        form.message.data = existing_invitation.message
        form.event_time.data = existing_invitation.event_time
        form.dress_code.data = existing_invitation.dress_code
        form.additional_info.data = existing_invitation.additional_info
    
    return render_template('create_invitation.html', booking=booking, form=form)


@app.route('/invitation/<token>')
def invitation_preview(token):
    """Preview invitation (for host to see)"""
    invitation = Invitation.query.filter_by(unique_token=token).first_or_404()
    booking = invitation.booking
    venue = booking.venue
    hall = booking.selected_hall
    
    total_invited = InvitedGuest.query.filter_by(invitation_id=invitation.id).count()
    attending = InvitedGuest.query.filter_by(invitation_id=invitation.id, rsvp_status='attending').count()
    not_attending = InvitedGuest.query.filter_by(invitation_id=invitation.id, rsvp_status='not_attending').count()
    pending = total_invited - attending - not_attending
    
    total_attending = db.session.query(
        db.func.sum(InvitedGuest.plus_one + 1)
    ).filter(
        InvitedGuest.invitation_id == invitation.id,
        InvitedGuest.rsvp_status == 'attending'
    ).scalar() or 0
    
    guests = InvitedGuest.query.filter_by(invitation_id=invitation.id).order_by(InvitedGuest.created_at.desc()).all()
    invitation_link = url_for('guest_rsvp_page', token=token, _external=True)
    
    return render_template('invitation_preview.html', 
                         invitation=invitation,
                         booking=booking,
                         venue=venue,
                         hall=hall,
                         invitation_link=invitation_link,
                         total_invited=total_invited,
                         attending=attending,
                         not_attending=not_attending,
                         pending=pending,
                         total_attending=total_attending,
                         guests=guests)


@app.route('/rsvp/<token>', methods=['GET', 'POST'])
def guest_rsvp_page(token):
    """Public RSVP page for guests"""
    invitation = Invitation.query.filter_by(unique_token=token).first_or_404()
    booking = invitation.booking
    venue = booking.venue
    hall = booking.selected_hall
    
    form = GuestRSVPForm()
    
    if form.validate_on_submit():
        # Create new RSVP (always create new entry, no duplicates check since no email)
        guest = InvitedGuest(
            invitation_id=invitation.id,
            name=form.name.data,
            email=None,  # Email removed from form
            phone=form.phone.data,
            plus_one=form.plus_one.data,
            rsvp_status=form.rsvp_status.data,
            dietary_restrictions=form.dietary_restrictions.data,
            message_to_host=form.message_to_host.data,
            responded_at=datetime.utcnow()
        )
        db.session.add(guest)
        
        db.session.commit()
        
        if form.rsvp_status.data == 'attending':
            flash('Thank you for confirming your attendance! We look forward to seeing you! 🎉', 'success')
        else:
            flash('Thank you for your response. We will miss you! 💔', 'info')
        
        return redirect(url_for('rsvp_confirmation', token=token))
    
    return render_template('guest_rsvp_page.html', 
                         invitation=invitation,
                         booking=booking,
                         venue=venue,
                         hall=hall,
                         form=form)


@app.route('/rsvp/<token>/confirmation')
def rsvp_confirmation(token):
    """Thank you page after RSVP submission"""
    invitation = Invitation.query.filter_by(unique_token=token).first_or_404()
    booking = invitation.booking
    venue = booking.venue
    
    return render_template('rsvp_confirmation.html', 
                         invitation=invitation,
                         booking=booking,
                         venue=venue)

if __name__ == '__main__':
    create_tables()
    app.run(debug=True)
